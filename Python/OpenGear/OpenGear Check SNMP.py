import pandas as pd
import requests
import json
import subprocess
import re
import pyodbc
from getpass import getpass
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from concurrent.futures import ThreadPoolExecutor
import time

# Reading file with hosts
def read_file():
    # Reading the config file
    while True:
        file = "hosts.txt" # input("Enter host file name to read: ")
        try:
            with open(file, "r", newline="") as host_file:
                host = host_file.readlines()
                host = [line.rstrip("\n\r") for line in host]  # Removing the \n
                return host
        except FileNotFoundError:
            print(f"No such file or directory: {file}")

def write_log(file_name):
    # Writing the log
    log_name = file_name.split(".")[0] + "_Logs.txt"
    with open(log_name, "a+") as write:
        for line in log_file:
            write.write(line + "\n")
    print(f"Log file: {log_name}")

# API Session
def api_sessions(og_url, og_username, og_password, api_ver):
    requests.packages.urllib3.disable_warnings()
    token = False
    api_session = '/api/%s/sessions' % api_ver

    data = {'username': og_username, 'password': og_password}

    # Authenticating
    try:
        session = requests.post(og_url + api_session, data=json.dumps(data), verify=False)
        session.raise_for_status()
    except requests.exceptions.RequestException as e:
        print("\nAccess to %s as %s failed" % (og_url, og_username))
        print(e)
        log_file.append(str(e))
        return

    if json.loads(session.text)['state'] == 'authenticated':
        token = json.loads(session.text)['session']
        message = f"Authentication successful to {og_url}"
        print(message)
        log_file.append(message)
    else:
        message = f"Authentication to {og_url} failed!"
        print(message)
        og_file.append(message)
    return token


# API GET
def api_get(host, og_username, og_password , hd):
    # API information
    api_ver = 'v2'
    og_url = "https://%s" % host
    token = api_sessions(og_url, og_username, og_password, api_ver)

    if token:
        headers = {'Authorization': 'Token ' + token}
    else:
        return
    
    # Get info
    get_snmpd = '/api/%s/services/snmpd' % api_ver
    get_firewall = '/api/%s/firewall/zones' % api_ver

    # Retrieving JSON for snmp
    try:
        r_snmp = requests.get(og_url + get_snmpd, headers=headers, verify=False)
        r_snmp.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(e)
    j_snmp = json.loads(r_snmp.text)

    # Retrieving JSON for firewall
    try:
        r_fw = requests.get(og_url + get_firewall, headers=headers, verify=False)
        r_fw.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(e)
    j_fw = json.loads(r_fw.text)

    # Info for excel output
    info_dict = {"Device": [host]}

    for line in hd:
        if line == "Device":
            continue
        # Firewall
        elif str(line) == "fw_intfs":
            try:
                info_dict["fw_intfs"] = [j_fw["firewall_zones"][0]["physifs"]]
            except:
                info_dict["fw_intfs"] = "Not found!"
        elif str(line) == "fw_services":
            try:
                info_dict["fw_services"] = [j_fw["firewall_zones"][0]["address_filters"][2]["services"]]
            except:
                info_dict["fw_services"] = "Not found!"
        elif str(line) == "fw_src":
            try:
                info_dict["fw_src"] = [j_fw["firewall_zones"][0]["address_filters"][2]["source_address"]]
            except:
                info_dict["fw_src"] = "Not found!"
        # SNMP
        else:
            try:
                info_dict[line] = j_snmp["snmpd"][line]
            except:
                info_dict[line] = "Not found!"

    # Fill missing values to ensure all lists have the same length
    max_length = max(len(lst) for lst in info_dict.values() if isinstance(lst, list))
    for key, value in info_dict.items():
        if isinstance(value, list) and len(value) < max_length:
            info_dict[key].extend([None] * (max_length - len(value)))
    new_row = pd.DataFrame(info_dict)
    return new_row


# Formatting the excel file
def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)

# Main loop
def main(host_file, file_name, sheet_name):
    global log_file
    log_file = []

    # Adding the header
    hd = [
        "Device",
        "priv_use_plaintext",
        "security_name",
        "protocol",
        "auth_password",
        "priv_protocol",
        "auth_protocol",
        "priv_password",
        "auth_use_plaintext",
        "enable_secure_snmp",
        "engine_id",
        "port",
        "enabled",
        "security_level",
        "enable_legacy_versions",
        "fw_intfs",
        "fw_services",
        "fw_src"
    ]
    df = pd.DataFrame(columns=hd)

    og_username = "a-deoliwx2"#input("Username: ")
    og_password = getpass()

    # Checking 20 devices per time
    with ThreadPoolExecutor(max_workers=50) as executor:
        future_to_device = {executor.submit(api_get, host, og_username, og_password, hd): host for host in host_file}

        # Process each result as they complete
        for future in future_to_device:
            new_row = future.result()
            try:
                if not new_row.empty:
                    df = pd.concat([df, new_row], ignore_index=True)
            except:
                continue

    df.to_excel(file_name, sheet_name=sheet_name, index=False)
    write_log(file_name)

if __name__ == "__main__":
    start_time = time.perf_counter()
    host_file = read_file()
    file_name = str(input("Enter the output file name: ")) + ".xlsx"
    sheet_name = "SNMP"
    main(host_file, file_name, sheet_name)
    excel_formatting(file_name, sheet_name)
    end_time = time.perf_counter()
    elapsed_time = end_time - start_time
    print(f"Total time taken: {elapsed_time:.4f} seconds")
    print("Ending program...")
