import pandas as pd
import requests
import json
import subprocess
import re
import pyodbc
from getpass import getpass
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


# Handle Singular or Plural values
def sop(num, stg):
	if num == 1:
		return("%d %s" % (num, stg)) # singular
	else:
		return("%d %s" % (num, stg + 's')) # plural

# Human Friendly Seconds
def hfs(secs):
	# Calculate days, hours, mins and seconds
	m, s = divmod(secs, 60)
	h, m = divmod(m, 60)
	d, h = divmod(h, 24)

	if d < 1:
		if h < 1:
			if m < 1:
				return("%s" % (sop(s,"second")))
			else:
				return("%s and %s" % (sop(m,"minute"), sop(s,"second")))
		else:
			return("%s, %s and %s" % (sop(h,"hour"), sop(m,"minute"), sop(s,"second")))
	else:
		return("%s, %s, %s and %s" % (sop(d,"day"), sop(h,"hour"), sop(m,"minute"), sop(s,"second")))

# Read CMDB from Database
def read_cmdb():
    server = 'xxxxxxx'
    database = 'xxxxx'
    username = input("DB Username: ")
    password = getpass()
    Authentication = 'ActiveDirectoryPassword'
    driver = '{ODBC Driver 17 for SQL Server}'

    conn = pyodbc.connect('DRIVER=' + driver +
                          ';SERVER=' + server +
                          ';PORT=1433;DATABASE=' + database +
                          ';UID=' + username +
                          ';PWD=' + password +
                          ';AUTHENTICATION=' + Authentication
                          )
    print(f"Authentication to Database...")
    cursor = conn.cursor()
    cursor.execute(
        "SELECT name, model_id_dv , install_status_dv , location_dv , location_region FROM snow.cmdb_ci WHERE sys_class_name = 'cmdb_ci_outofband_device' AND (manufacturer_dv LIKE '%open gear%' OR manufacturer_dv LIKE '%opengear%')")
    cmdb = {}
    for row in cursor.fetchall():
        cmdb[row[0]] = {"Model": row[1], "Install_Status": row[2], "CMDB_Location": row[3], "CMDB_Region": row[4]}
    return cmdb

# Formatting the excel file
def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)

# Main loop
def main(cmdb, file_name, sheet_name):
    API_VER = 'v3.7'
    API_SESSIONS = '/api/%s/sessions' % API_VER
    API_NODES_ENROLLED = '/api/%s/nodes?config:status=Enrolled' % API_VER

    lhaddress = "lighthouse.domain.com"
    lhusername = input("Lighthouse Username: ")
    lhpassword = getpass()

    print("Authenticating to %s as %s" % (lhaddress, lhusername))

    LH5_URL = "https://%s" % lhaddress

    requests.packages.urllib3.disable_warnings()

    data = {'username': lhusername, 'password': lhpassword}
    #Authenticating
    try:
        r = requests.post(LH5_URL + API_SESSIONS, data=json.dumps(data), verify=False)
        r.raise_for_status()
    except requests.exceptions.RequestException as e:
        print("\nAuthentication to %s as %s failed" % (LH5_URL, lhusername))
        print(e)
        exit(1)
    if json.loads(r.text)['state'] == 'authenticated':
        token = json.loads(r.text)['session']
        print("Authentication successful")
    else:
        print("Authentication failed - exiting")
        exit(99)

    headers = {'Authorization': 'Token ' + token}
    #Retreiving JSON
    try:
        print("Retrieving JSON...")
        r = requests.get(LH5_URL + API_NODES_ENROLLED, headers=headers, verify=False)
        r.raise_for_status()
        print("Retrieved successfully")
    except requests.exceptions.RequestException as e:
        print(e)
        exit(1)
    j = json.loads(r.text)
    #Header
    df = pd.DataFrame(columns=["Device", "IP", "Mac_address", "Firmware_version", "Serial_number", "Enrollment_bundle",
                               "Connection_status", "Cell_Health", "Model", "PTR", "CMDB_model", "CMDB_install_Status",
                               "CMDB_Location", "CMDB_Region", "Last_status_change"])
    #Looping tru json
    node_count = len(j["nodes"])
    print(f"Found {node_count} nodes")
    for node in j["nodes"]:
        lsc = hfs(node['runtime_status']['change_delta']) #checking the Last status change with hfs
        intf = []
        cell_health = node["cellhealth_runtime_status"]["status"]
        for row in node["interfaces"]:
            if "ipv4_addr" in row:
                if "192.168." not in row["ipv4_addr"] and "wan" not in row["name"] and "Cellular Modem" not in row[
                    "name"]:
                    intf.append(row["name"] + " " + row["ipv4_addr"])
                else:
                    continue
        #Running nslookup to check dns entry
        try:
            node_ip = re.search(r"\d+\.\d+\.\d+\.\d+", str(intf)).group(0)
        except:
            node_ip = "No"
        x = str(subprocess.run(f"nslookup {node_ip}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True))
        try:
            dns = re.search(f"{node_ip}.+\d+\.\d+\.\d+\.\d+.+Name:\s+(\S+\.com)", x).group(1)
        except:
            dns = "Not found"

        #CMDB information

        for ci in cmdb:
            if ci.lower() == node["name"].lower():
                cmdb_model = cmdb[ci]["Model"]
                install_status = cmdb[ci]["Install_Status"]
                cmdb_location = cmdb[ci]["CMDB_Location"]
                cmdb_region = cmdb[ci]["CMDB_Region"]
                break
            else:
                cmdb_model = "Not found"
                install_status = "Not found"
                cmdb_location = "Not found"
                cmdb_region = "Not found"

        new_row = pd.DataFrame({
            "Device": [node["name"].lower()],
            "IP": [intf],
            "Mac_address": node["mac_address"],
            "Firmware_version": node["firmware_version"],
            "Serial_number": str(node["serial_number"]),
            "Enrollment_bundle": node["enrollment_bundle"],
            "Connection_status": node["runtime_status"]["connection_status"],
            "Cell_Health": cell_health,
            "Model": node["model"],
            "PTR": dns,
            "CMDB_model": cmdb_model,
            "CMDB_install_Status": install_status,
            "CMDB_Location": cmdb_location,
            "CMDB_Region": cmdb_region,
            "Last_status_change": lsc
        })
        df = pd.concat([df, new_row], ignore_index=True)
        # Counting down
        print(f"\rReading Device Number: {node_count}  ", end="")
        node_count -= 1
    df.to_excel(file_name, sheet_name=sheet_name, index=False)
    print(f"\nFile name is {file_name}")

if __name__ == "__main__":
    cmdb = read_cmdb()
    file_name = str(input("Enter the output file name: ")) + ".xlsx"
    sheet_name = "Lighthouse Inventory"
    main(cmdb, file_name, sheet_name)
    excel_formatting(file_name, sheet_name)
    print("Ending program...")
