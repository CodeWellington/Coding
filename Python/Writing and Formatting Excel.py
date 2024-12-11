import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

def excel_writing(file_name, sheet_name):
    # example
    devices = {"SW1": "10.10.10.10",
               "RTR1": "20.20.20.20"}

    # Create a DataFrame header example
    data = pd.DataFrame(columns=["Device", "IP"])

    # Append new data
    for dev in devices:
        new_row = pd.DataFrame({"Device": [dev], "IP": [devices[dev]]})
        data = pd.concat([data, new_row], ignore_index=True)

    # Write the combined data to the Excel file
    data.to_excel(file_name, sheet_name=sheet_name, index=False)

def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)


if __name__ == "__main__":
    file_name = str(input("Enter the output file name: ")) + ".xlsx"
    sheet_name = "via python"
    excel_writing(file_name, sheet_name)
    excel_formatting(file_name, sheet_name)

