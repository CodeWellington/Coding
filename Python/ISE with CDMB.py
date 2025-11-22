
import csv
import json
import oracledb
import signal
import ssl
import sys
import time
import maskpass
import pyodbc
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Reading data from CMDB - bisdw
def read_cmdb():
    server = 'xxxxx.database.windows.net'
    database = 'xxxx'
    username = input("Enter your CMDB DB Username: ")
    password = maskpass.askpass(prompt="Enter your CMDB DB password: ")
    Authentication = 'ActiveDirectoryPassword'
    driver = '{ODBC Driver 17 for SQL Server}'

    try:
        conn = pyodbc.connect('DRIVER=' + driver +
                              ';SERVER=' + server +
                              ';PORT=1433;DATABASE=' + database +
                              ';UID=' + username +
                              ';PWD=' + password +
                              ';AUTHENTICATION=' + Authentication
                              )
    except Exception as e:
        print(f"Error: {e}")
        quit()

    print(f"Authentication to Database...")

    # Query to database
    cursor = conn.cursor()
    cursor.execute(
        """SELECT
            a.MAC,
            a.Asset_Name,
            a.CMDB_Status,
            a.CMDB_URL,
            a.CMDB_Class,
            c.location_dv AS CMDB_Location

        FROM it.Asset a -- from table it.asset
        JOIN snow.cmdb_ci c -- location is on snow.cmdb_ci table
            ON a.Asset_Name = c.name  -- adding the location based on name
        WHERE a.CMDB_URL IS NOT NULL
          AND a.MAC IS NOT NULL
          AND LOWER(CMDB_Status) NOT LIKE ('%retired%')""")
    cmdb = {}
    print("Querying the DB...")

    # Building the cmdb dict
    for row in cursor.fetchall():
        cmdb[row[0].upper()] = {
            "Asset_Name": row[1],
            "CMDB_Status": row[2],
            "CMDB_URL": row[3],
            "CMDB_Class": row[4],
            "CMDB_Location": row[5]
        }
    return cmdb

# Access dataconnect and check against cmdb
def data_connect(dc_query, file_name, cmdb):
    hostname = "xxxxxx.com"
    password = maskpass.askpass(prompt="Enter Dataconnet password: ")
    ISE_DC_USERNAME = "dataconnect"
    ISE_DC_PORT = 2484  # Data Connect port
    ISE_DC_SID = "cpm10"  # Data Connect service name identifier

    ssl_context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)

    # Ignoring Cert error
    ssl_context.check_hostname = False  # required before setting verify_mode == ssl.CERT_NONE
    ssl_context.verify_mode = ssl.CERT_NONE  # any cert is accepted; validation errors are ignored

    # oracle parameters
    params = oracledb.ConnectParams(
        protocol="tcps",  # tcp "secure" with TLS
        host=hostname,  # hostname or IP address of database host machine
        port=ISE_DC_PORT,  # Oracle Default: 1521
        service_name=ISE_DC_SID,
        user=ISE_DC_USERNAME,
        password=password,
        retry_count=3,  # connection attempts retries before being terminated. Default: 0
        retry_delay=3,  # seconds to wait before a new connection attempt. Default: 0
        ssl_context=ssl_context,  # an SSLContext object which is used for connecting to the database using TLS
        ssl_server_dn_match=False,
        # boolean indicating if the server certificate distinguished name (DN) should be matched. Default: True
    )

    try:
        with oracledb.connect(params=params) as connection:
            with connection.cursor() as cursor:
                cursor.execute(dc_query)

                headers = [f"{i[0]}".lower() for i in cursor.description]
                headers.append("cmdb_asset")
                headers.append("cmdb_status")
                headers.append("cmdb_class")
                headers.append("cmdb_location")
                headers.append("cmdb_url")

                df = pd.DataFrame(columns=headers)
                new_row = {}
                unique_macs = set()  # Use a set for uniqueness
                print(f"Reading rows ...")

                while True:
                    rows = cursor.fetchmany()  # use default Cursor.arraysize
                    if not rows: # It retrieves chunks of information and stop when finishes
                        break



                    # loop thru each row
                    for value in rows:  # tuples

                        # Getting the unique MACS
                        mac = value[1]
                        if mac:
                            unique_macs.add(mac)

                        for index in range(len(value)):
                            if headers[index] == "calling_station_id":

                                if cmdb.get(value[index]):
                                    new_row["cmdb_asset"] = cmdb[value[index]]["Asset_Name"]
                                    new_row["cmdb_status"] = cmdb[value[index]]["CMDB_Status"]
                                    new_row["cmdb_class"] = cmdb[value[index]]["CMDB_Class"]
                                    new_row["cmdb_location"] = cmdb[value[index]]["CMDB_Location"]
                                    new_row["cmdb_url"] = cmdb[value[index]]["CMDB_URL"]
                                else:
                                    new_row["cmdb_asset"] = "Not Found"
                                    new_row["cmdb_status"] = "Not Found"
                                    new_row["cmdb_class"] = "Not Found"
                                    new_row["cmdb_location"] = "Not Found"
                                    new_row["cmdb_url"] = "Not Found"

                            new_row[headers[index]] = [value[index]]
                        new_row = pd.DataFrame(new_row)
                        df = pd.concat([df, new_row], ignore_index=True)


                return df, unique_macs

    except oracledb.Error as e:
        print(f"Oracle Error: {e}")
        quit()
    except Exception as e:
        print(f"Error: {e}")
        quit()

#Formatting the excel
def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName=sheet_name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)

def unique_mac_info(unique_macs):
    new_row = {}
    headers = ["mac", "cmdb_asset", "cmdb_status", "cmdb_class", "cmdb_location", "cmdb_url"]

    df = pd.DataFrame(columns=headers)

    for mac in unique_macs:
        if cmdb.get(mac):
            new_row["mac"] = [mac]
            new_row["cmdb_asset"] = [cmdb[mac]["Asset_Name"]]
            new_row["cmdb_status"] = cmdb[mac]["CMDB_Status"]
            new_row["cmdb_class"] = cmdb[mac]["CMDB_Class"]
            new_row["cmdb_location"] = cmdb[mac]["CMDB_Location"]
            new_row["cmdb_url"] = cmdb[mac]["CMDB_URL"]
        else:
            new_row["mac"] = [mac]
            new_row["cmdb_asset"] = "Not Found"
            new_row["cmdb_status"] = "Not Found"
            new_row["cmdb_class"] = "Not Found"
            new_row["cmdb_location"] = "Not Found"
            new_row["cmdb_url"] = "Not Found"
            pass
        new_row = pd.DataFrame(new_row)
        df = pd.concat([df, new_row], ignore_index=True)
    return df


if __name__ == "__main__":

    start_time = time.time()
    file_name = str(input("Enter the output file name: ")) + ".xlsx"
    dc_query = """
    SELECT
        TO_CHAR(timestamp, 'YYYY-MM-DD HH24:MI:SS') AS timestamp, 
        calling_station_id,
        device_name as sw_name,
        nas_port_id as sw_port,
        ise_node,
        policy_set_name, 
        access_service,
        authentication_method AS auth_method,
        authorization_rule AS authz_rule, 
        authorization_profiles AS authz_profiles,
        endpoint_profile
    FROM radius_authentications

    WHERE timestamp > SYSDATE - INTERVAL '1' DAY -- CHANGE THIS LIKE TO FILTER +OR- DAYS
      AND username NOT IN ('DUMMY', 'ISEDUMMY')
      AND calling_station_id IS NOT NULL
      AND access_service NOT in ('Wireless 802.1x')
      AND device_name LIKE 'P2824%' -- CHANGE THIS LIKE TO FILTER, DB TOO BIG
    ORDER BY timestamp DESC
    """
    cmdb = read_cmdb()
    df_ise, unique_macs = data_connect(dc_query, file_name, cmdb)
    df_macs = unique_mac_info(unique_macs)

    # Create one Excel file with two sheets
    with pd.ExcelWriter(file_name, engine="openpyxl") as writer:
        df_ise.to_excel(writer, sheet_name="ISE", index=False)
        df_macs.to_excel(writer, sheet_name="MACS", index=False)

    excel_formatting(file_name, "ISE")
    excel_formatting(file_name, "MACS")


    print(f"Total run time: {'{0:.3f}'.format(time.time() - start_time)} seconds", file=sys.stderr)
