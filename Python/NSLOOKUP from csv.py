from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import subprocess
import pandas as pd
import re
import csv
import os

# Reading CSV
with open(r"MAC.csv", "r") as file:
    info = csv.DictReader(file)
    output = [row for row in info]


# Running nslookup to check dns entry
def nslookup(ip):
    x = str(
        subprocess.run(f"nslookup {str(ip)}", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True))
    try:
        dns = re.search(f"{ip}.+\d+\.\d+\.\d+\.\d+.+Name:\s+(\S+\.com)", x).group(1)
    except:
        dns = "Not found"

    return dns


def excel_formatting(file_name, sheet_name, table_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName=table_name, ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)


def main(file_name):
    # Header
    df = pd.DataFrame(columns=["Device", "Interface", "VLAN", "MAC", "IP", "DNS", "Vendor", "Description", "Neighbour",
                               "Neighbour_Model",
                               "Neighbour_Capabilities"])

    # Getting all IPs and Unique VLANs
    ip_list = []
    vlan_list = []
    for node in output:
        ip_list.append(node["IP"])
        if node["VLAN"] not in vlan_list:
            vlan_list.append(node["VLAN"])

    # nslookup for each IP
    with ThreadPoolExecutor(max_workers=80) as executor:
        future_to_device = {executor.submit(nslookup, ip): ip for ip in ip_list}

        ip_dns = {}
        # Process each result as they complete
        for future in as_completed(future_to_device):
            host = future_to_device[future]

            try:
                result = future.result()
                ip_dns[host] = result
            except Exception as e:
                print(f"Error processing {host}: {e}")

    # Create a new workbook if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Add sheets based on VLANs
    for vlan in vlan_list:
        vlan_df = df[df['VLAN'] == vlan]

        for node in output:
            if node["VLAN"] == vlan:
                try:
                    new_row = pd.DataFrame({
                        "Device": [node["Device"]],
                        "Interface": node["Interface"],
                        "VLAN": node["VLAN"],
                        "MAC": node["MAC"],
                        "IP": node["IP"],
                        "DNS": ip_dns.get(node['IP']),
                        "Vendor": node["Vendor"],
                        "Description": node["Description"],
                        "Neighbour": node["Neighbour"],
                        "Neighbour_Model": node["Neighbour_Model"],
                        "Neighbour_Capabilities": node["Neighbour_Capabilities"]})
                    vlan_df = pd.concat([vlan_df, new_row], ignore_index=True)
                except Exception as e:
                    print(f"Error processing {node['IP']}: {e}")

        sheet_name = f"VLAN_{vlan}"
        table_name = f"Table_{vlan}"
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
            vlan_df.to_excel(writer, sheet_name=sheet_name, index=False)

        excel_formatting(file_name, sheet_name, table_name)


if __name__ == "__main__":
    file_name = str(input("Enter the output file name: ")) + ".xlsx"
    main(file_name)
    print("Ending program...")
