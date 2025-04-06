import textfsm
import json
import time
import paramiko
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

# Reading file with hosts
def read_file():
    # Reading the config file
    while True:
        file = input("Enter the txt host file name to read: ")
        try:
            with open(file, "r", newline="") as host_file:
                host = host_file.readlines()
                host = [line.rstrip("\n\r") for line in host]  # Removing the \n\r
                return host
        except FileNotFoundError:
            print(f"No such file or directory: {file}")


# Access Device
def ssh(host):
    ssh_client = paramiko.SSHClient()  #ssh client from paramiko
    ssh_client.set_missing_host_key_policy(paramiko.AutoAddPolicy()) #Helps when the key is unkown
    ssh_client.connect(username='ansible', password='ansible', hostname=host, port=22, timeout=60,  allow_agent=False, look_for_keys=False)
    transport = ssh_client.get_transport()
    session = transport.open_session()
    session.invoke_shell()
    session.send('\r')
    time.sleep(2)
    return session

# Reading textfsm file
def read_textfsm():
    template_file = open('basic.textfsm')
    textfsm_parser = textfsm.TextFSM(template_file)
    template_file.close()
    return textfsm_parser


def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)


def main(file_name, sheet_name):
    info_dict = {}

    # Reading textfsm file
    textfsm_parser = read_textfsm()

    # Host to access
    host = "100.65.0.1"

    # Accessing the device
    commands = ssh(host)

    # Sending commands
    commands.send(b"terminal length 0\n")
    commands.recv(99999999)
    commands.send(b"show ip interface brief\n")
    time.sleep(1)

    # output of the commands.send
    log = str(commands.recv(99999999).decode("utf-8"))

    # Parsing the data
    parsed_data = textfsm_parser.ParseText(log)

    # Formating the data
    parsed_output = [dict(zip(textfsm_parser.header, row)) for row in parsed_data]
    print(json.dumps(parsed_output, indent=4))

    # Header
    df = pd.DataFrame(columns=textfsm_parser.header)

    # writhing each row to csv without the header
    for row in parsed_output:
        new_row = {}
        for k, v in row.items():
            new_row[k] = [v]
        new_row = pd.DataFrame(new_row)
        df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(file_name, sheet_name=sheet_name, index=False)

file_name = str(input("Enter the output file name: ")) + ".xlsx"
sheet_name = "Inventory"
main(file_name,sheet_name)
excel_formatting(file_name, sheet_name)




