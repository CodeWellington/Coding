
import csv
import json
import oracledb
import signal
import ssl
import sys
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font
from openpyxl.worksheet.table import Table, TableStyleInfo


def excel_formatting(file_name, sheet_name):
    # Load the workbook and select the sheet
    wb = load_workbook(file_name)
    ws = wb[sheet_name]

    # Define the border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Apply the border to all cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Change the header font color to white
    header_font = Font(color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font

    # AutoFit column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Add table formatting
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Save the workbook
    wb.save(file_name)



start_time = time.time()

# Merge settings from 1) CLI args, 2) environment variables
hostname = input("Enter the hostname: ")
password = "mypass"
ISE_DC_USERNAME = "myuser"
ISE_DC_PORT = 2484  # Data Connect port
ISE_DC_SID = "cpm10"  # Data Connect service name identifier


ssl_context = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)

# Ignoring Cert error
ssl_context.check_hostname = False  # required before setting verify_mode == ssl.CERT_NONE
ssl_context.verify_mode = ssl.CERT_NONE  # any cert is accepted; validation errors are ignored

params = oracledb.ConnectParams(
    protocol="tcps",  # tcp "secure" with TLS
    host=hostname,  # hostname or IP address of database host machine
    port=ISE_DC_PORT,  # Oracle Default: 1521
    service_name=ISE_DC_SID,
    user=ISE_DC_USERNAME,
    password=password,
    retry_count=3,  # connection attempts retries before being terminated. Default: 0
    retry_delay=3,  # seconds to wait before a new connection attempt. Default: 0
    ssl_context=ssl_context,  # an SSLContext object which is used for connecting to the database using TLS
    ssl_server_dn_match=False,  # boolean indicating if the server certificate distinguished name (DN) should be matched. Default: True
)

def main(file_name,sheet_name, query):
    try:
        with oracledb.connect(params=params) as connection:
            with connection.cursor() as cursor:

                print(f"SQL query:\n-----\n{query}\n-----")
                cursor.execute(query)

                headers = [f"{i[0]}".lower() for i in cursor.description]
                print("HEADERS")
                print(headers)
                df = pd.DataFrame(columns=headers)
                new_row = {}
                print(f"Reading rows ...")
                while True:
                    rows = cursor.fetchmany()  # use default Cursor.arraysize
                    if not rows:
                        df.to_excel(file_name, sheet_name=sheet_name, index=False)
                        break
                    for value in rows: # tuples
                        for index in range(len(value)):
                            new_row[headers[index]] = [value[index]]
                            print(f"\rReading rows ...", end="")
                        new_row = pd.DataFrame(new_row)
                        df = pd.concat([df, new_row], ignore_index=True)

    except oracledb.Error as e:
        print(f"Oracle Error: {e}")
    except Exception as e:
        print(f"Error: {e}")

file_name = str(input("Enter the output file name: ")) + ".xlsx"
sheet_name = "ISE"
query = """
SELECT
    TO_CHAR(timestamp, 'YYYY-MM-DD HH24:MI:SS') AS timestamp, 
    calling_station_id,
    username,
    device_name,
    NAS_PORT_ID,
    ise_node,
    policy_set_name, 
    access_service,
    authentication_method AS auth_method,
    authentication_protocol AS auth_protocol,
    authorization_rule AS authz_rule, 
    authorization_profiles AS authz_profiles,
    response_time
FROM radius_authentications
   WHERE timestamp > sysdate - INTERVAL '10' DAY 
   AND username != 'DUMMY'
ORDER BY timestamp DESC -- most recent records
"""
main(file_name,sheet_name, query)
excel_formatting(file_name, sheet_name)

print(f"Total run time: {'{0:.3f}'.format(time.time() - start_time)} seconds", file=sys.stderr)
